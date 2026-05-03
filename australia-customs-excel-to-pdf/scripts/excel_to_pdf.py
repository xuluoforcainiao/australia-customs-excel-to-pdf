import sys
import os
import zipfile
import openpyxl
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import HexColor

def register_font():
    font_paths = [
        'C:/Windows/Fonts/msyh.ttc',
        'C:/Windows/Fonts/simsun.ttc',
        '/System/Library/Fonts/PingFang.ttc',
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('CJKFont', fp, subfontIndex=0))
                return 'CJKFont'
            except:
                continue
    return 'Helvetica'

def extract_images(xlsx_path, output_dir):
    images = []
    try:
        z = zipfile.ZipFile(xlsx_path)
        idx = 1
        for name in z.namelist():
            if name.startswith('xl/media/') and not name.endswith('/'):
                data = z.read(name)
                ext = name.split('.')[-1]
                out_path = os.path.join(output_dir, f'image{idx}.{ext}')
                with open(out_path, 'wb') as f:
                    f.write(data)
                images.append(out_path)
                idx += 1
    except Exception as e:
        print(f'Warning: could not extract images: {e}')
    return images

def find_cell(ws, keyword, start_row=1, end_row=20):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and keyword.lower() in cell.value.lower():
                return cell
    return None

def get_val(ws, row, col):
    cell = ws.cell(row=row, column=col)
    return cell.value

def generate_pdf(excel_path, output_pdf):
    font_name = register_font()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    if len(wb.sheetnames) > 1:
        ws = wb['IV'] if 'IV' in wb.sheetnames else wb.active

    # Extract images
    tmp_dir = os.path.dirname(output_pdf)
    images = extract_images(excel_path, tmp_dir)

    width, height = landscape(A4)
    c = canvas.Canvas(output_pdf, pagesize=landscape(A4))

    def draw_text(x, y, text, font=font_name, size=10, align='left', color=None):
        if color:
            c.setFillColor(color)
        else:
            c.setFillColorRGB(0, 0, 0)
        c.setFont(font, size)
        if align == 'center':
            c.drawCentredString(x, y, str(text) if text is not None else '')
        elif align == 'right':
            c.drawRightString(x, y, str(text) if text is not None else '')
        else:
            c.drawString(x, y, str(text) if text is not None else '')

    red_color = HexColor('#CC0000')
    blue_color = HexColor('#0000CC')

    left_margin = 2*cm
    right_margin = 2*cm
    page_width = width - left_margin - right_margin

    # Outer border
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(1.5)
    c.rect(left_margin, 2*cm, page_width, height - 4*cm, fill=0, stroke=1)

    # Company header
    start_y = height - 2.8*cm
    company_name = ws['D1'].value or 'GUANGZHOU HAYONEX LOGISTICS CO.,LTD--固定无需修改'
    company_addr = ws['D2'].value or '24TH FLOOR, JIN HUI BUILDING,NO.123,JIE FANG NAN ROAD,YUE XIU DISTRICT,GUANGZHOU,CHINA.-----固定无需修改'
    # Clean up garbled chars if any
    company_name = company_name.replace('\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd', '\u56fa\u5b9a\u65e0\u9700\u4fee\u6539').replace('\ufffd', '')
    company_addr = company_addr.replace('\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd\ufffd', '\u56fa\u5b9a\u65e0\u9700\u4fee\u6539').replace('\ufffd', '')

    draw_text(width/2, start_y, company_name, font_name, 12, align='center', color=red_color)
    start_y -= 0.6*cm
    draw_text(width/2, start_y, company_addr, font_name, 9, align='center', color=red_color)
    start_y -= 0.35*cm
    c.setLineWidth(0.5)
    c.line(left_margin + 0.5*cm, start_y, width - left_margin - 0.5*cm, start_y)

    # Title
    start_y -= 0.6*cm
    draw_text(width/2, start_y, 'COMMERCIAL INVOICE', font_name, 13, align='center')
    start_y -= 0.35*cm
    c.line(left_margin + 0.5*cm, start_y, width - left_margin - 0.5*cm, start_y)
    start_y -= 0.5*cm

    # Consignee
    consignee = ws['E4'].value or ''
    draw_text(left_margin + 0.5*cm, start_y, 'Consignee :', font_name, 10)
    draw_text(left_margin + 3*cm, start_y, consignee, font_name, 10, color=red_color)

    # Right side info
    right_label_x = width/2 + 2*cm
    right_val_x = width - left_margin - 1*cm
    info_y = start_y
    info_items = [
        ('DATE:', str(ws['J4'].value) if ws['J4'].value else ''),
        ('INVOICE NO.:', str(ws['J5'].value) if ws['J5'].value else ''),
        ('TRADE TERM:', str(ws['J6'].value) if ws['J6'].value else ''),
        ('PORT OF LOADING:', str(ws['J7'].value) if ws['J7'].value else ''),
        ('DESTINATION:', str(ws['J8'].value) if ws['J8'].value else ''),
    ]
    for label, val in info_items:
        draw_text(right_label_x, info_y, label, font_name, 10)
        c.setLineWidth(0.5)
        c.line(right_label_x + 3.5*cm, info_y - 0.1*cm, right_val_x, info_y - 0.1*cm)
        draw_text(right_val_x, info_y, val, font_name, 10, align='right')
        info_y -= 0.5*cm

    start_y = min(start_y, info_y) - 0.7*cm

    # Table
    table_left = left_margin + 0.5*cm
    table_right = width - left_margin - 0.5*cm
    table_width = table_right - table_left

    cols = [3.5*cm, 2.5*cm, 2.5*cm, 4.5*cm, 3*cm, 3.5*cm, 4*cm]
    scale = table_width / sum(cols)
    cols = [w * scale for w in cols]
    col_x = [table_left]
    for w in cols[:-1]:
        col_x.append(col_x[-1] + w)

    header_height = 0.6*cm
    c.setFillColorRGB(0.7, 0.7, 0.7)
    c.rect(table_left, start_y - header_height, table_width, header_height, fill=1, stroke=1)
    c.setFillColorRGB(0, 0, 0)

    headers = ['BUYER P. O. NO.', 'Package', 'Quantity', 'DESCRIPTION OF GOODS', "Products' Material", 'UNIT PRICE (AUD)', 'TOTAL AMOUNT (AUD)']
    for x, w, h in zip(col_x, cols, headers):
        c.setFont(font_name, 9)
        c.drawCentredString(x + w/2, start_y - 0.4*cm, h)

    data_y = start_y - header_height - 0.45*cm
    data_row = [
        ws['D10'].value, ws['E10'].value, ws['F10'].value,
        ws['G10'].value, ws['H10'].value, ws['I10'].value, ws['J10'].value
    ]
    for x, w, val in zip(col_x, cols, data_row):
        c.setFont(font_name, 10)
        c.drawCentredString(x + w/2, data_y, str(val) if val is not None else '')

    total_y = data_y - 0.5*cm
    grand_total = [
        ws['D11'].value, ws['E11'].value, ws['F11'].value,
        None, None, None, ws['J11'].value
    ]
    for x, w, val in zip(col_x, cols, grand_total):
        if val is not None:
            c.setFont(font_name, 10)
            c.drawCentredString(x + w/2, total_y, str(val))

    table_bottom = total_y - 0.3*cm
    c.setLineWidth(0.5)
    c.rect(table_left, table_bottom, table_width, start_y - table_bottom, fill=0, stroke=1)
    c.line(table_left, start_y - header_height, table_right, start_y - header_height)
    c.line(table_left, data_y + 0.3*cm, table_right, data_y + 0.3*cm)
    c.line(table_left, total_y + 0.3*cm, table_right, total_y + 0.3*cm)
    for x in col_x[1:]:
        c.line(x, table_bottom, x, start_y)

    # WEB
    web_y = table_bottom - 0.5*cm
    draw_text(table_left + 0.1*cm, web_y, 'WEB:', font_name, 9)
    url = ws['E12'].value or ''
    c.setFillColor(blue_color)
    c.setFont(font_name, 8)
    c.drawString(table_left + 1.5*cm, web_y, url)
    c.setFillColorRGB(0, 0, 0)

    # Images
    if images:
        img_y = web_y - 5*cm
        if img_y > 2*cm:
            for img_path in images[:1]:
                try:
                    img = ImageReader(img_path)
                    c.drawImage(img, table_left + 0.5*cm, img_y, width=5.5*cm, height=4.8*cm, preserveAspectRatio=True)
                except Exception as e:
                    print(f'Warning: could not draw image: {e}')

    c.save()
    print(f'PDF generated: {output_pdf}')

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python excel_to_pdf.py <input.xlsx> <output.pdf>')
        sys.exit(1)
    generate_pdf(sys.argv[1], sys.argv[2])
